import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import sys
import argparse
import re
from utils import find_ticker_dir, get_latest_financial_data, ExcelStyles, setup_logging

logger = setup_logging("generate_3statement")

def build_3statement_model(ticker_data, ticker_dir):
    wb = openpyxl.Workbook()
    
    # 反復計算有効化（3表の循環参照解決に必須）
    calc_pr = CalcProperties(iterate=True, refMode='A1', iterateCount=100, iterateDelta=0.001)
    wb.properties.calcPr = calc_pr
    
    ws = wb.active
    ws.title = "財務3表モデル"
    ws.views.sheetView[0].showGridLines = True
    
    # 共通スタイルのインポート
    styles = ExcelStyles()
    
    font_family = styles.font_family
    title_font = styles.title_font
    header_font = styles.header_font
    section_font = styles.section_font
    data_font = styles.data_font
    input_font = styles.input_font
    bold_data_font = styles.bold_data_font
    primary_fill = styles.primary_fill
    section_fill = styles.section_fill
    highlight_fill = styles.highlight_fill
    thin_border_side = styles.thin_border_side
    thin_border = styles.thin_border
    
    is_jpy = ticker_data["currency"] == "JPY"
    unit_str = "十億円" if is_jpy else "百万米ドル"
    div_factor = 1e9 if is_jpy else 1e6
    
    # 実績データの単位変換
    rev_act = ticker_data["revenue"] / div_factor
    ebit_act = ticker_data["ebit"] / div_factor
    da_act = ticker_data["depreciation"] / div_factor
    tax_act = ticker_data["tax_provision"] / div_factor
    debt_act = ticker_data["total_debt"] / div_factor
    cash_act = ticker_data["cash"] / div_factor
    
    # タイトル
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{ticker_data['name']} ({ticker_data['ticker']}) - 財務3表連動モデル"
    ws["A1"].font = title_font
    ws["A1"].fill = primary_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Inputsシートを作成して前提条件を配置
    ws_inputs = wb.create_sheet(title="前提条件")
    ws_inputs.views.sheetView[0].showGridLines = True

    # タイトル行 (Inputs)
    ws_inputs.merge_cells("A1:C1")
    ws_inputs["A1"] = f"{ticker_data['name']} ({ticker_data['ticker']}) - 3表モデル 前提条件"
    ws_inputs["A1"].font = title_font
    ws_inputs["A1"].fill = primary_fill
    ws_inputs["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_inputs.row_dimensions[1].height = 40

    ws_inputs.cell(row=3, column=1, value="項目").font = header_font
    ws_inputs.cell(row=3, column=1).fill = primary_fill
    ws_inputs.cell(row=3, column=2, value="設定値").font = header_font
    ws_inputs.cell(row=3, column=2).fill = primary_fill
    ws_inputs.cell(row=3, column=3, value="情報源 / 備考").font = header_font
    ws_inputs.cell(row=3, column=3).fill = primary_fill
    ws_inputs.row_dimensions[3].height = 25

    inputs_data = [
        # Revenue Growth
        ("売上高成長率 FY25E", 0.12, "0.0%", "[想定]"),
        ("売上高成長率 FY26E", 0.10, "0.0%", "[想定]"),
        ("売上高成長率 FY27E", 0.08, "0.0%", "[想定]"),
        ("売上高成長率 FY28E", 0.06, "0.0%", "[想定]"),
        ("売上高成長率 FY29E", 0.05, "0.0%", "[想定]"),
        ("Dummy", "", "", ""), # Row 9
        # COGS % of Rev
        ("売上原価率 FY25E", 0.64, "0.0%", "[想定]"),
        ("売上原価率 FY26E", 0.63, "0.0%", "[想定]"),
        ("売上原価率 FY27E", 0.62, "0.0%", "[想定]"),
        ("売上原価率 FY28E", 0.62, "0.0%", "[想定]"),
        ("売上原価率 FY29E", 0.62, "0.0%", "[想定]"),
        ("Dummy", "", "", ""), # Row 15
        # SG&A % of Rev
        ("販管費率 FY25E", 0.145, "0.0%", "[想定]"),
        ("販管費率 FY26E", 0.14, "0.0%", "[想定]"),
        ("販管費率 FY27E", 0.14, "0.0%", "[想定]"),
        ("販管費率 FY28E", 0.14, "0.0%", "[想定]"),
        ("販管費率 FY29E", 0.14, "0.0%", "[想定]"),
        ("Dummy", "", "", ""), # Row 21
        # D&A % of Rev
        ("減価償却費率 FY25E", 0.10, "0.0%", "[想定]"),
        ("減価償却費率 FY26E", 0.10, "0.0%", "[想定]"),
        ("減価償却費率 FY27E", 0.095, "0.0%", "[想定]"),
        ("減価償却費率 FY28E", 0.09, "0.0%", "[想定]"),
        ("減価償却費率 FY29E", 0.09, "0.0%", "[想定]"),
        ("Dummy", "", "", ""), # Row 27
        # Debt Interest Rate
        ("有利子負債利率", 0.025, "0.0%", "[想定] 負債コスト"),
        ("Dummy", "", "", ""), # Row 29
        # AR/Inv/AP % of Rev
        ("売上債権比率 (%/売上高)", 0.12, "0.0%", "[想定]"),
        ("棚卸資産比率 (%/売上高)", 0.17, "0.0%", "[想定]"),
        ("仕入債務比率 (%/売上高)", 0.08, "0.0%", "[想定]"),
    ]

    inputs_map = {}

    for idx, (label, val, fmt, src) in enumerate(inputs_data):
        row = 4 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=label).font = data_font
        if val != "":
            cell_val = ws_inputs.cell(row=row, column=2, value=val)
            cell_val.font = input_font
            cell_val.number_format = fmt
            cell_val.alignment = Alignment(horizontal="right")
            inputs_map[label] = f"前提条件!$B${row}"
        ws_inputs.cell(row=row, column=3, value=src).font = data_font
        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # CapEx Projections
    ws_inputs.cell(row=38, column=1, value="設備投資計画").font = bold_data_font
    ws_inputs.cell(row=39, column=1, value="年度").font = header_font
    ws_inputs.cell(row=39, column=1).fill = primary_fill
    ws_inputs.cell(row=39, column=2, value="設備投資額").font = header_font
    ws_inputs.cell(row=39, column=2).fill = primary_fill
    ws_inputs.cell(row=39, column=3, value="情報源 / 備考").font = header_font
    ws_inputs.cell(row=39, column=3).fill = primary_fill
    ws_inputs.row_dimensions[39].height = 20

    capex_inputs = [
        ("FY25E", -300.0, "[想定]"),
        ("FY26E", -320.0, "[想定]"),
        ("FY27E", -310.0, "[想定]"),
        ("FY28E", -300.0, "[想定]"),
        ("FY29E", -300.0, "[想定]"),
    ]
    for idx, (yr, val, src) in enumerate(capex_inputs):
        row = 40 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=yr).font = data_font
        cell_val = ws_inputs.cell(row=row, column=2, value=val)
        cell_val.font = input_font
        cell_val.number_format = "#,##0.0"
        cell_val.alignment = Alignment(horizontal="right")
        ws_inputs.cell(row=row, column=3, value=src).font = data_font
        inputs_map[f"設備投資額 {yr}"] = f"前提条件!$B${row}"
        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # Debt Drawdown / (Repayment)
    ws_inputs.cell(row=46, column=1, value="有利子負債の調達・返済計画").font = bold_data_font
    ws_inputs.cell(row=47, column=1, value="年度").font = header_font
    ws_inputs.cell(row=47, column=1).fill = primary_fill
    ws_inputs.cell(row=47, column=2, value="調達・返済額").font = header_font
    ws_inputs.cell(row=47, column=2).fill = primary_fill
    ws_inputs.cell(row=47, column=3, value="情報源 / 備考").font = header_font
    ws_inputs.cell(row=47, column=3).fill = primary_fill
    ws_inputs.row_dimensions[47].height = 20

    debt_inputs = [
        ("FY25E", 50.0, "[想定]"),
        ("FY26E", -20.0, "[想定]"),
        ("FY27E", -30.0, "[想定]"),
        ("FY28E", -40.0, "[想定]"),
        ("FY29E", -50.0, "[想定]"),
    ]
    for idx, (yr, val, src) in enumerate(debt_inputs):
        row = 48 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=yr).font = data_font
        cell_val = ws_inputs.cell(row=row, column=2, value=val)
        cell_val.font = input_font
        cell_val.number_format = "#,##0.0"
        cell_val.alignment = Alignment(horizontal="right")
        ws_inputs.cell(row=row, column=3, value=src).font = data_font
        inputs_map[f"有利子負債の増減 {yr}"] = f"前提条件!$B${row}"
        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # Inputsシートの列幅調整
    for col in ws_inputs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_inputs.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # ヘッダー
    headers = [f"財務3表連動モデル ({unit_str})", "FY24A", "FY25E", "FY26E", "FY27E", "FY28E", "FY29E"]
    for col_idx, header in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=header)
        c.font = header_font
        c.fill = primary_fill
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 28
    
    # 1. 損益計算書 (Income Statement)
    ws["A5"] = "I. 損益計算書"
    ws["A5"].font = section_font
    ws.merge_cells("A5:G5")
    ws["A5"].fill = section_fill
    
    tax_rate_ltm = tax_act / ebit_act if ebit_act > 0 else 0.306
    
    is_rows = [
        ("売上高", [rev_act, "=B{row}*(1+{Inputs:売上高成長率 FY25E})", "=C{row}*(1+{Inputs:売上高成長率 FY26E})", "=D{row}*(1+{Inputs:売上高成長率 FY27E})", "=E{row}*(1+{Inputs:売上高成長率 FY28E})", "=F{row}*(1+{Inputs:売上高成長率 FY29E})"]),
        ("売上高成長率", ["", "=(C{売上高}-B{売上高})/B{売上高}", "=(D{売上高}-C{売上高})/C{売上高}", "=(E{売上高}-D{売上高})/D{売上高}", "=(F{売上高}-E{売上高})/E{売上高}", "=(G{売上高}-F{売上高})/F{売上高}"]),
        ("売上原価", [rev_act * 0.65, "=C{売上高}*{Inputs:売上原価率 FY25E}", "=D{売上高}*{Inputs:売上原価率 FY26E}", "=E{売上高}*{Inputs:売上原価率 FY27E}", "=F{売上高}*{Inputs:売上原価率 FY28E}", "=G{売上高}*{Inputs:売上原価率 FY29E}"]),
        ("売上総利益", ["=B{売上高}-B{売上原価}", "=C{売上高}-C{売上原価}", "=D{売上高}-D{売上原価}", "=E{売上高}-E{売上原価}", "=F{売上高}-F{売上原価}", "=G{売上高}-G{売上原価}"]),
        ("販売費及び一般管理費 (販管費)", [rev_act * 0.15, "=C{売上高}*{Inputs:販管費率 FY25E}", "=D{売上高}*{Inputs:販管費率 FY26E}", "=E{売上高}*{Inputs:販管費率 FY27E}", "=F{売上高}*{Inputs:販管費率 FY28E}", "=G{売上高}*{Inputs:販管費率 FY29E}"]),
        ("EBITDA", ["=B{売上総利益}-B{販売費及び一般管理費 (販管費)}", "=C{売上総利益}-C{販売費及び一般管理費 (販管費)}", "=D{売上総利益}-D{販売費及び一般管理費 (販管費)}", "=E{売上総利益}-E{販売費及び一般管理費 (販管費)}", "=F{売上総利益}-F{販売費及び一般管理費 (販管費)}", "=G{売上総利益}-G{販売費及び一般管理費 (販管費)}"]),
        ("減価償却費", [da_act, "=C{売上高}*{Inputs:減価償却費率 FY25E}", "=D{売上高}*{Inputs:減価償却費率 FY26E}", "=E{売上高}*{Inputs:減価償却費率 FY27E}", "=F{売上高}*{Inputs:減価償却費率 FY28E}", "=G{売上高}*{Inputs:減価償却費率 FY29E}"]),
        ("営業利益 (EBIT)", ["=B{EBITDA}-B{減価償却費}", "=C{EBITDA}-C{減価償却費}", "=D{EBITDA}-D{減価償却費}", "=E{EBITDA}-E{減価償却費}", "=F{EBITDA}-F{減価償却費}", "=G{EBITDA}-G{減価償却費}"]),
        ("支払利息", [15.0, "=AVERAGE(B{有利子負債},C{有利子負債})*{Inputs:有利子負債利率}", "=AVERAGE(C{有利子負債},D{有利子負債})*{Inputs:有利子負債利率}", "=AVERAGE(D{有利子負債},E{有利子負債})*{Inputs:有利子負債利率}", "=AVERAGE(E{有利子負債},F{有利子負債})*{Inputs:有利子負債利率}", "=AVERAGE(F{有利子負債},G{有利子負債})*{Inputs:有利子負債利率}"]),
        ("税引前当期純利益", ["=B{営業利益 (EBIT)}-B{支払利息}", "=C{営業利益 (EBIT)}-C{支払利息}", "=D{営業利益 (EBIT)}-D{支払利息}", "=E{営業利益 (EBIT)}-E{支払利息}", "=F{営業利益 (EBIT)}-F{支払利息}", "=G{営業利益 (EBIT)}-G{支払利息}"]),
        ("法人税等", [tax_act, f"=C{{税引前当期純利益}}*{tax_rate_ltm:.4f}", f"=D{{税引前当期純利益}}*{tax_rate_ltm:.4f}", f"=E{{税引前当期純利益}}*{tax_rate_ltm:.4f}", f"=F{{税引前当期純利益}}*{tax_rate_ltm:.4f}", f"=G{{税引前当期純利益}}*{tax_rate_ltm:.4f}"]),
        ("当期純利益", ["=B{税引前当期純利益}-B{法人税等}", "=C{税引前当期純利益}-C{法人税等}", "=D{税引前当期純利益}-D{法人税等}", "=E{税引前当期純利益}-E{法人税等}", "=F{税引前当期純利益}-F{法人税等}", "=G{税引前当期純利益}-G{法人税等}"])
    ]
    
    # 2. 貸借対照表 (Balance Sheet)
    bs_rows = [
        ("資産の部", ["", "", "", "", "", ""]),
        ("現金及び現金同等物", [cash_act, "=B{現金及び現金同等物の期末残高}", "=C{現金及び現金同等物の期末残高}", "=D{現金及び現金同等物の期末残高}", "=E{現金及び現金同等物の期末残高}", "=F{現金及び現金同等物の期末残高}"]),
        ("売上債権", [rev_act * 0.12, "=C{売上高}*{Inputs:売上債権比率 (%/売上高)}", "=D{売上高}*{Inputs:売上債権比率 (%/売上高)}", "=E{売上高}*{Inputs:売上債権比率 (%/売上高)}", "=F{売上高}*{Inputs:売上債権比率 (%/売上高)}", "=G{売上高}*{Inputs:売上債権比率 (%/売上高)}"]),
        ("棚卸資産", [rev_act * 0.18, "=C{売上高}*{Inputs:棚卸資産比率 (%/売上高)}", "=D{売上高}*{Inputs:棚卸資産比率 (%/売上高)}", "=E{売上高}*{Inputs:棚卸資産比率 (%/売上高)}", "=F{売上高}*{Inputs:棚卸資産比率 (%/売上高)}", "=G{売上高}*{Inputs:棚卸資産比率 (%/売上高)}"]),
        ("有形固定資産 (純額)", [1200.0, "=B{row}+C{設備投資額 (CapEx)}-C{減価償却費}", "=C{row}+D{設備投資額 (CapEx)}-D{減価償却費}", "=D{row}+E{設備投資額 (CapEx)}-E{減価償却費}", "=E{row}+F{設備投資額 (CapEx)}-F{減価償却費}", "=F{row}+G{設備投資額 (CapEx)}-G{減価償却費}"]),
        ("資産合計", ["=SUM(B{現金及び現金同等物}:B{有形固定資産 (純額)})", "=SUM(C{現金及び現金同等物}:C{有形固定資産 (純額)})", "=SUM(D{現金及び現金同等物}:D{有形固定資産 (純額)})", "=SUM(E{現金及び現金同等物}:E{有形固定資産 (純額)})", "=SUM(F{現金及び現金同等物}:F{有形固定資産 (純額)})", "=SUM(G{現金及び現金同等物}:G{有形固定資産 (純額)})"]),
        
        ("負債及び純資産の部", ["", "", "", "", "", ""]),
        ("仕入債務", [rev_act * 0.08, "=C{売上高}*{Inputs:仕入債務比率 (%/売上高)}", "=D{売上高}*{Inputs:仕入債務比率 (%/売上高)}", "=E{売上高}*{Inputs:仕入債務比率 (%/売上高)}", "=F{売上高}*{Inputs:仕入債務比率 (%/売上高)}", "=G{売上高}*{Inputs:仕入債務比率 (%/売上高)}"]),
        ("有利子負債", [debt_act, "=B{row}+C{有利子負債の増減}", "=C{row}+D{有利子負債の増減}", "=D{row}+E{有利子負債の増減}", "=E{row}+F{有利子負債の増減}", "=F{row}+G{有利子負債の増減}"]),
        ("負債合計", ["=B{仕入債務}+B{有利子負債}", "=C{仕入債務}+C{有利子負債}", "=D{仕入債務}+D{有利子負債}", "=E{仕入債務}+E{有利子負債}", "=F{仕入債務}+F{有利子負債}", "=G{仕入債務}+G{有利子負債}"]),
        
        ("資本金", [800.0, "=B{row}", "=C{row}", "=D{row}", "=E{row}", "=F{row}"]),
        ("利益剰余金", [300.0, "=B{row}+C{当期純利益}", "=C{row}+D{当期純利益}", "=D{row}+E{当期純利益}", "=E{row}+F{当期純利益}", "=F{row}+G{当期純利益}"]),
        ("純資産合計", ["=B{資本金}+B{利益剰余金}", "=C{資本金}+C{利益剰余金}", "=D{資本金}+D{利益剰余金}", "=E{資本金}+E{利益剰余金}", "=F{資本金}+F{利益剰余金}", "=G{資本金}+G{利益剰余金}"]),
        ("負債純資産合計", ["=B{負債合計}+B{純資産合計}", "=C{負債合計}+C{純資産合計}", "=D{負債合計}+D{純資産合計}", "=E{負債合計}+E{純資産合計}", "=F{負債合計}+F{純資産合計}", "=G{負債合計}+G{純資産合計}"]),
        
        ("バランス確認 (不一致額)", ["=B{資産合計}-B{負債純資産合計}", "=C{資産合計}-C{負債純資産合計}", "=D{資産合計}-D{負債純資産合計}", "=E{資産合計}-E{負債純資産合計}", "=F{資産合計}-F{負債純資産合計}", "=G{資産合計}-G{負債純資産合計}"])
    ]
    
    # 3. キャッシュ・フロー計算書 (Cash Flow Statement)
    cf_rows = [
        ("当期純利益 (CF調整用)", ["=B{当期純利益}", "=C{当期純利益}", "=D{当期純利益}", "=E{当期純利益}", "=F{当期純利益}", "=G{当期純利益}"]),
        ("加算：減価償却費", ["=B{減価償却費}", "=C{減価償却費}", "=D{減価償却費}", "=E{減価償却費}", "=F{減価償却費}", "=G{減価償却費}"]),
        ("減算：売上債権の増減", ["", "=-(C{売上債権}-B{売上債権})", "=-(D{売上債権}-C{売上債権})", "=-(E{売上債権}-D{売上債権})", "=-(F{売上債権}-E{売上債権})", "=-(G{売上債権}-F{売上債権})"]),
        ("減算：棚卸資産の増減", ["", "=-(C{棚卸資産}-B{棚卸資産})", "=-(D{棚卸資産}-C{棚卸資産})", "=-(E{棚卸資産}-D{棚卸資産})", "=-(F{棚卸資産}-E{棚卸資産})", "=-(G{棚卸資産}-F{棚卸資産})"]),
        ("加算：仕入債務の増減", ["", "=C{仕入債務}-B{仕入債務}", "=D{仕入債務}-C{仕入債務}", "=E{仕入債務}-D{仕入債務}", "=F{仕入債務}-E{仕入債務}", "=G{仕入債務}-F{仕入債務}"]),
        ("営業活動によるキャッシュ・フロー", ["=SUM(B{当期純利益 (CF調整用)}:B{加算：仕入債務の増減})", "=SUM(C{当期純利益 (CF調整用)}:C{加算：仕入債務の増減})", "=SUM(D{当期純利益 (CF調整用)}:D{加算：仕入債務の増減})", "=SUM(E{当期純利益 (CF調整用)}:E{加算：仕入債務の増減})", "=SUM(F{当期純利益 (CF調整用)}:F{加算：仕入債務の増減})", "=SUM(G{当期純利益 (CF調整用)}:G{加算：仕入債務の増減})"]),
        
        ("設備投資額 (CapEx)", [-225.6, "={Inputs:設備投資額 FY25E}", "={Inputs:設備投資額 FY26E}", "={Inputs:設備投資額 FY27E}", "={Inputs:設備投資額 FY28E}", "={Inputs:設備投資額 FY29E}"]),
        ("投資活動によるキャッシュ・フロー", ["=B{設備投資額 (CapEx)}", "=C{設備投資額 (CapEx)}", "=D{設備投資額 (CapEx)}", "=E{設備投資額 (CapEx)}", "=F{設備投資額 (CapEx)}", "=G{設備投資額 (CapEx)}"]),
        
        ("有利子負債の増減", [-324.3, "={Inputs:有利子負債の増減 FY25E}", "={Inputs:有利子負債の増減 FY26E}", "={Inputs:有利子負債の増減 FY27E}", "={Inputs:有利子負債の増減 FY28E}", "={Inputs:有利子負債の増減 FY29E}"]),
        ("財務活動によるキャッシュ・フロー", ["=B{有利子負債の増減}", "=C{有利子負債の増減}", "=D{有利子負債の増減}", "=E{有利子負債の増減}", "=F{有利子負債の増減}", "=G{有利子負債の増減}"]),
        
        ("現金及び現金同等物の増減額", ["=B{営業活動によるキャッシュ・フロー}+B{投資活動によるキャッシュ・フロー}+B{財務活動によるキャッシュ・フロー}", "=C{営業活動によるキャッシュ・フロー}+C{投資活動によるキャッシュ・フロー}+C{財務活動によるキャッシュ・フロー}", "=D{営業活動によるキャッシュ・フロー}+D{投資活動によるキャッシュ・フロー}+D{財務活動によるキャッシュ・フロー}", "=E{営業活動によるキャッシュ・フロー}+E{投資活動によるキャッシュ・フロー}+E{財務活動によるキャッシュ・フロー}", "=F{営業活動によるキャッシュ・フロー}+F{投資活動によるキャッシュ・フロー}+F{財務活動によるキャッシュ・フロー}", "=G{営業活動によるキャッシュ・フロー}+G{投資活動によるキャッシュ・フロー}+G{財務活動によるキャッシュ・フロー}"]),
        ("現金及び現金同等物の期首残高", [187.6, "=B{現金及び現金同等物の期末残高}", "=C{現金及び現金同等物の期末残高}", "=D{現金及び現金同等物の期末残高}", "=E{現金及び現金同等物の期末残高}", "=F{現金及び現金同等物の期末残高}"]),
        ("現金及び現金同等物の期末残高", ["=B{現金及び現金同等物の期首残高}+B{現金及び現金同等物の増減額}", "=C{現金及び現金同等物の期首残高}+C{現金及び現金同等物の増減額}", "=D{現金及び現金同等物の期首残高}+D{現金及び現金同等物の増減額}", "=E{現金及び現金同等物の期首残高}+E{現金及び現金同等物の増減額}", "=F{現金及び現金同等物の期首残高}+F{現金及び現金同等物の増減額}", "=G{現金及び現金同等物の期首残高}+G{現金及び現金同等物の増減額}"])
    ]
    
    # 行番号を動的に事前確定する
    row_map = {}
    
    # 1. IS の行番号確定
    current_row = 6
    for label, _ in is_rows:
        row_map[label] = current_row
        current_row += 1
        
    # 2. BS の行番号確定
    title_row_bs = current_row + 1 # タイトル行 (II. BALANCE SHEET)
    current_row = title_row_bs + 1 # ASSETSラベルの行
    for label, _ in bs_rows:
        row_map[label] = current_row
        current_row += 1
        
    # 3. CF の行番号確定
    title_row_cf = current_row + 1 # タイトル行 (III. CASH FLOW STATEMENT)
    current_row = title_row_cf + 1
    for label, _ in cf_rows:
        row_map[label] = current_row
        current_row += 1
 
    # テンプレート数式を解決するヘルパー関数
    def resolve_formula(template, col_idx, current_row_num):
        if not isinstance(template, str) or not template.startswith("="):
            return template
            
        col_letter = get_column_letter(col_idx)
        prev_col_letter = get_column_letter(col_idx - 1) if col_idx > 2 else ""
        
        formula = template
        
        # {Inputs:xxx} の置換
        def replace_inputs(match):
            key = match.group(1)
            return inputs_map.get(key, f"前提条件!$B$4")
            
        formula = re.sub(r'\{Inputs:([^}]+)\}', replace_inputs, formula)
        
        # {row} 自体の置換 (現在の行)
        formula = formula.replace("{row}", str(current_row_num))
        
        # {prev_col_letter} や {col_letter} を補完するためにプレースホルダー内の置換を行う
        # {prev:xxx} の置換
        def replace_prev(match):
            key = match.group(1)
            row = row_map.get(key)
            if row:
                return f"{prev_col_letter}{row}"
            return f"{prev_col_letter}1"
            
        formula = re.sub(r'\{prev:([^}]+)\}', replace_prev, formula)
        
        # {xxx} の置換 (現在の列 of 特定行)
        def replace_curr(match):
            key = match.group(1)
            row = row_map.get(key)
            if row:
                return f"{col_letter}{row}"
            return f"{col_letter}1"
            
        formula = re.sub(r'\{([^}]+)\}', replace_curr, formula)
        
        return formula

    # データを流し込む関数
    def populate_section(rows, start_row):
        for idx, (label, vals) in enumerate(rows):
            r = start_row + idx
            ws.row_dimensions[r].height = 22
            
            is_bold = label in ["売上高", "売上総利益", "EBITDA", "当期純利益", "資産合計", "負債合計", "純資産合計", "負債純資産合計", "営業活動によるキャッシュ・フロー", "投資活動によるキャッシュ・フロー", "財務活動によるキャッシュ・フロー", "現金及び現金同等物の増減額", "現金及び現金同等物の期末残高", "バランス確認 (不一致額)"]
            
            cell_label = ws.cell(row=r, column=1, value=label)
            cell_label.font = bold_data_font if is_bold else data_font
            
            if label == "バランス確認 (不一致額)":
                cell_label.fill = highlight_fill
                cell_label.font = Font(name=font_family, size=11, bold=True, color="047857")
                
            for col_idx, val in enumerate(vals, 2):
                cell = ws.cell(row=r, column=col_idx)
                cell.alignment = Alignment(horizontal="right")
                
                # 数式テンプレートの動的解決
                resolved_val = resolve_formula(val, col_idx, r)
                
                if str(resolved_val).startswith("="):
                    cell.value = resolved_val
                    cell.font = bold_data_font if is_bold else data_font
                else:
                    cell.value = resolved_val
                    cell.font = input_font if (col_idx == 2 or label in ["設備投資額 (CapEx)", "有利子負債の増減", "売上高"]) else data_font
                
                if "%" in label or "成長率" in label:
                    cell.number_format = '0.0%'
                elif label == "バランス確認 (不一致額)":
                    cell.number_format = '0.00'
                    cell.fill = highlight_fill
                else:
                    cell.number_format = '#,##0.0'
                    
                if is_bold:
                    cell.border = Border(top=thin_border_side, bottom=Side(style='double' if "合計" in label or "バランス" in label or "期末残高" in label or "当期純利益" in label else 'thin', color='1B263B'))
                    
        return start_row + len(rows)
        
    # I. ISを描画
    last_row = populate_section(is_rows, 6)
    
    # II. BSを描画
    ws.cell(row=title_row_bs, column=1, value="II. 貸借対照表").font = section_font
    ws.merge_cells(start_row=title_row_bs, start_column=1, end_row=title_row_bs, end_column=7)
    ws.cell(row=title_row_bs, column=1).fill = section_fill
    last_row = populate_section(bs_rows, title_row_bs + 1)
    
    # III. CFを描画
    ws.cell(row=title_row_cf, column=1, value="III. キャッシュ・フロー計算書").font = section_font
    ws.merge_cells(start_row=title_row_cf, start_column=1, end_row=title_row_cf, end_column=7)
    ws.cell(row=title_row_cf, column=1).fill = section_fill
    last_row = populate_section(cf_rows, title_row_cf + 1)
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    target_path = os.path.join(ticker_dir, "analysis")
    os.makedirs(target_path, exist_ok=True)
    out_file = os.path.join(target_path, f"3statement_{ticker_data['ticker']}.xlsx")
    wb.save(out_file)
    wb.close()
    logger.info(f"Successfully generated 3-Statement model: {out_file}")

def main():
    parser = argparse.ArgumentParser(description="Generate 3-statement financial model")
    parser.add_argument("ticker", type=str, help="Stock ticker")
    parser.add_argument("--outdir", type=str, default="./out", help="Base output directory")
    args = parser.parse_args()
    
    ticker_str = args.ticker.strip()
    ticker_dir = find_ticker_dir(args.outdir, ticker_str)
    
    try:
        ticker_data = get_latest_financial_data(ticker_dir, ticker_str)
    except FileNotFoundError as e:
        logger.error(str(e))
        sys.exit(1)
        
    build_3statement_model(ticker_data, ticker_dir)

if __name__ == "__main__":
    main()
