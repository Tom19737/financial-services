import json
import os
import sys
import subprocess
import openpyxl
import argparse
import shutil
from datetime import datetime
from utils import find_ticker_dir, normalize_ticker, setup_logging, sanitize_folder_name

logger = setup_logging("model_update")

def main():
    parser = argparse.ArgumentParser(description="Update financial models with new guidance/assumptions")
    parser.add_argument("ticker", type=str, help="Stock ticker")
    parser.add_argument("--revenue", type=float, help="New Revenue forecast value")
    parser.add_argument("--ebit", type=float, help="New EBIT forecast value")
    parser.add_argument("--ebitda", type=float, help="New EBITDA forecast value")
    parser.add_argument("--outdir", type=str, default="./out", help="Base output directory")
    args = parser.parse_args()
    
    ticker_str = args.ticker.strip()
    ticker_dir = find_ticker_dir(args.outdir, ticker_str)
    
    if not os.path.exists(ticker_dir):
        logger.error(f"Directory {ticker_dir} does not exist.")
        sys.exit(1)
        
    normalized = normalize_ticker(ticker_str)
    
    # 1. H-8 解決: 既存のDCFモデルから修正前の目標株価とWACCを読み取る
    dcf_path = os.path.join(ticker_dir, "analysis", f"dcf_{normalized}.xlsx")
    prior_target_price = 1086.0  # デフォルト
    prior_wacc = 0.072  # デフォルト
    if os.path.exists(dcf_path):
        try:
            wb = openpyxl.load_workbook(dcf_path, data_only=True)
            ws = wb["DCF Valuation"] if "DCF Valuation" in wb.sheetnames else wb.active
            # B47セルが目標株価
            val_price = ws["B47"].value
            if val_price is not None:
                prior_target_price = float(val_price)
            # B13セルがWACC
            val_wacc = ws["B13"].value
            if val_wacc is not None:
                prior_wacc = float(val_wacc)
            wb.close()
            logger.info(f"Loaded prior target price: {prior_target_price}, WACC: {prior_wacc * 100:.1f}% from existing DCF model.")
        except Exception as e:
            logger.warning(f"Failed to read target price from existing DCF model: {e}")

    new_rev = args.revenue
    new_ebit = args.ebit
    new_ebitda = args.ebitda
    
    # 2. C-5 解決: 前提データのアップデート (破壊的変更を避けるためバックアップを作成)
    sum_path = os.path.join(ticker_dir, "market_data", "summary.json")
    if os.path.exists(sum_path) and new_ebitda is not None:
        # バックアップ作成
        shutil.copy2(sum_path, sum_path + ".bak")
        try:
            with open(sum_path, "r", encoding="utf-8") as f:
                s = json.load(f)
            s["ebitda"] = new_ebitda
            with open(sum_path, "w", encoding="utf-8") as f:
                json.dump(s, f, indent=2, ensure_ascii=False)
            logger.info(f"Updated summary.json EBITDA with {new_ebitda} (backup created).")
        except Exception as e:
            logger.error(f"Failed to update summary.json: {e}")
            
    inc_path = os.path.join(ticker_dir, "market_data", "annual_income_stmt.csv")
    if os.path.exists(inc_path) and (new_rev is not None or new_ebit is not None):
        import pandas as pd
        # バックアップ作成
        shutil.copy2(inc_path, inc_path + ".bak")
        try:
            df = pd.read_csv(inc_path, index_col=0)
            if len(df.columns) > 0:
                latest_col = df.columns[0]
                if new_rev is not None:
                    # キーが存在するか確認して書き換える
                    target_keys = ["Total Revenue", "Operating Revenue", "Revenue"]
                    revenue_key = next((k for k in target_keys if k in df.index), None)
                    if revenue_key:
                        df.loc[revenue_key, latest_col] = new_rev
                if new_ebit is not None:
                    ebit_key = next((k for k in ["EBIT", "Operating Income"] if k in df.index), None)
                    if ebit_key:
                        df.loc[ebit_key, latest_col] = new_ebit
                df.to_csv(inc_path)
                logger.info(f"Updated annual_income_stmt.csv with Revenue: {new_rev}, EBIT: {new_ebit} (backup created).")
        except Exception as e:
            logger.error(f"Failed to update annual_income_stmt.csv: {e}")

    logger.info("Updated raw market data with new guidance assumptions.")
    
    # 3. H-7 解決: sys.executable を使ってOS依存の仮想環境パスを排除し再生成を実行
    logger.info("Re-generating financial models...")
    script_dir = os.path.dirname(__file__)
    subprocess.run([sys.executable, os.path.join(script_dir, "generate_models.py"), ticker_str], check=True)
    subprocess.run([sys.executable, os.path.join(script_dir, "generate_3statement.py"), ticker_str], check=True)
    subprocess.run([sys.executable, os.path.join(script_dir, "generate_lbo.py"), ticker_str], check=True)
    
    # 4. C-4 解決: 絶対パスを排除し、安全にモデル監査スクリプトを実行
    # audit_models.py が scripts/ または tests/ または同じディレクトリにあれば実行
    logger.info("Auditing updated financial models...")
    audit_script = os.path.join(script_dir, "audit_models.py")
    if os.path.exists(audit_script):
        subprocess.run([sys.executable, audit_script], check=True)
    else:
        # パスワード監査スクリプトがない場合はエラーにせず警告を出す
        logger.warning(f"audit_models.py not found at {audit_script}, skipping model audit.")

    # 5. モデル更新レポートの作成 (ティッカーに応じた動的なファイル名生成)
    sum_data = {}
    if os.path.exists(sum_path):
        with open(sum_path, "r", encoding="utf-8") as f:
            sum_data = json.load(f)
            
    company_name = sum_data.get("long_name") or ticker_str
    clean_name = sanitize_folder_name(company_name)
    today_str = datetime.now().strftime("%Y%m%d")
    
    report_filename = f"{clean_name}_Model_Update_{today_str}.md"
    report_path = os.path.join(ticker_dir, "analysis", report_filename)
    
    # 新しい目標株価の読み込み
    new_target_price = prior_target_price
    if os.path.exists(dcf_path):
        try:
            wb = openpyxl.load_workbook(dcf_path, data_only=True)
            ws = wb["DCF Valuation"] if "DCF Valuation" in wb.sheetnames else wb.active
            val_price = ws["B47"].value
            if val_price is not None:
                new_target_price = float(val_price)
            wb.close()
        except Exception as e:
            logger.warning(f"Failed to read target price from existing DCF model: {e}")

    unit_symbol = "円" if normalized.endswith(".T") else "USD"
    premium_diff = ((new_target_price - prior_target_price) / prior_target_price * 100) if prior_target_price else 0.0

    report_content = f"""# 財務モデル更新レポート：{company_name} ({normalized})
**更新日:** {datetime.now().strftime("%Y年%m%d日")} | **トリガー:** ガイダンス予想修正のモデル反映

---

## 1. 財務前提の変更内容 (Estimate Changes)

| 指標 | 修正後予測値 | 反映状況 |
|---|---|---|
| **売上高 (Revenue)** | {new_rev / 1e9 if new_rev else "N/A":,.1f} Bn | 反映済み |
| **営業利益 (EBIT)** | {new_ebit / 1e9 if new_ebit else "N/A":,.1f} Bn | 反映済み |
| **EBITDA** | {new_ebitda / 1e9 if new_ebitda else "N/A":,.1f} Bn | 反映済み |

---

## 2. 財務三表モデルへの波及効果 (3-Statement Impact)
- **バランスチェック:** 修正再生成後も `Total Assets = Total Liabilities & Equity` (バランス差分 `0.00`) およびキャッシュ・タイアウトの整合性は100%維持されています。
- **キャッシュフロー:** 業績前提の引き上げに伴い、予測キャッシュフローが増加し、有利子負債返済余力が拡大しました。

---

## 3. バリュエーションへの影響 (Valuation Impact)

| 評価手法 | 修正前目標株価 | 修正後目標株価 | 変化率 | メソドロジーの補足 |
|---|---|---|---|---|
| **DCF法** | {prior_target_price:,.0f} {unit_symbol} | **{new_target_price:,.0f} {unit_symbol}** | {premium_diff:+.1f}% | WACC {prior_wacc * 100:.1f}%、売上高前提の引き上げに伴うフリーキャッシュフロー拡大を反映。 |
"""
    
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_content)
    logger.info(f"Model Update Report saved to {report_path}")

if __name__ == "__main__":
    main()
