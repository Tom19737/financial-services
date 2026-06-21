import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
import os
import sys
import json
import argparse
import subprocess
from utils import normalize_ticker, find_ticker_dir, get_latest_financial_data, ExcelStyles, setup_logging

logger = setup_logging("generate_models")

def parse_args():
    parser = argparse.ArgumentParser(description="Generate generic DCF and Comps models for a given stock ticker")
    parser.add_argument("ticker", type=str, help="Subject stock ticker (e.g. 7203, 285A, MSFT)")
    parser.add_argument("--peers", type=str, help="Comma-separated list of peer tickers for Comps (e.g. 7267,7201 or MU,WDC)")
    parser.add_argument("--outdir", type=str, default="./out", help="Base output directory")
    return parser.parse_args()

def fetch_peer_data_if_missing(peer_ticker, outdir):
    """ピア企業のデータが不足している場合、fetch_yfinance.py を自動実行して補う"""
    peer_dir = find_ticker_dir(outdir, peer_ticker)
    summary_path = os.path.join(peer_dir, "market_data", "summary.json")
    if not os.path.exists(summary_path):
        logger.info(f"Peer data for {peer_ticker} is missing. Fetching using fetch_yfinance.py...")
        try:
            script_path = os.path.join(os.path.dirname(__file__), "fetch_yfinance.py")
            subprocess.run([
                sys.executable, 
                script_path, 
                peer_ticker, 
                "--skip-fundamentals", 
                "--outdir", outdir
            ], check=True)
        except Exception as e:
            logger.error(f"Failed to fetch peer {peer_ticker}: {e}")

def create_comps_model(ticker_data, peers_list, outdir):
    """類似企業比較 (Comps) シートを構築する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "類似企業比較"
    ws.views.sheetView[0].showGridLines = True
    
    # 共通スタイルインポート
    styles = ExcelStyles()
    
    font_family = styles.font_family
    title_font = styles.title_font
    header_font = styles.header_font
    data_font = styles.data_font
    input_font = styles.input_font
    bold_data_font = styles.bold_data_font
    primary_fill = styles.primary_fill
    section_fill = styles.section_fill
    subject_fill = styles.subject_fill
    thin_border = styles.thin_border
    
    is_jpy = ticker_data["currency"] == "JPY"
    unit_str = "十億円" if is_jpy else "百万米ドル"
    div_factor = 1e9 if is_jpy else 1e6
    
    # タイトル
    ws.merge_cells("A1:K1")
    ws["A1"] = f"{ticker_data['name']} ({ticker_data['ticker']}) - 類似企業比較分析"
    ws["A1"].font = title_font
    ws["A1"].fill = primary_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # ヘッダー行
    headers = [
        "企業名", "ティッカー", "通貨", f"時価総額\n({unit_str})", f"企業価値 (EV)\n({unit_str})",
        f"売上高\n({unit_str})", "売上高成長率\n(前年比)", f"EBITDA\n({unit_str})", "EBITDAマージン",
        "EV / 売上高", "EV / EBITDA"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = primary_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 35
    
    # ピアデータの集計
    all_data = []
    # 1. ピア企業
    for peer in peers_list:
        fetch_peer_data_if_missing(peer, outdir)
        p_dir = find_ticker_dir(outdir, peer)
        try:
            p_data = get_latest_financial_data(p_dir, peer)
            all_data.append((p_data["name"], peer, p_data["currency"], p_data["market_cap"], p_data["revenue"], p_data["ebitda"]))
        except Exception as e:
            logger.warning(f"Skipping peer {peer} due to missing data: {e}")
        
    # 2. 対象企業 (最後に配置)
    all_data.append((f"{ticker_data['name']} (対象会社)", ticker_data["ticker"], ticker_data["currency"], ticker_data["market_cap"], ticker_data["revenue"], ticker_data["ebitda"]))
    
    start_row = 4
    for idx, (name, tk, curr, mc, rev, ebitda) in enumerate(all_data):
        row = start_row + idx
        ws.row_dimensions[row].height = 24
        
        ws.cell(row=row, column=1, value=name).font = bold_data_font if "対象会社" in name else data_font
        ws.cell(row=row, column=2, value=tk).font = input_font
        ws.cell(row=row, column=3, value=curr).font = input_font
        
        mc_val = mc / div_factor if mc else 0.0
        rev_val = rev / div_factor if rev else 0.0
        eb_val = ebitda / div_factor if ebitda else 0.0
        
        # 簡易EV計算
        ev_val = mc_val
        peer_dir = find_ticker_dir(outdir, tk)
        sum_path = os.path.join(peer_dir, "market_data", "summary.json")
        if os.path.exists(sum_path):
            with open(sum_path, "r", encoding="utf-8") as f:
                s = json.load(f)
                if "market_cap" in s and s["market_cap"]:
                    ev_info = s.get("enterprise_value") or s.get("market_cap")
                    ev_val = float(ev_info) / div_factor
                if tk == ticker_data["ticker"]:
                    ev_val = mc_val + (ticker_data["total_debt"] - ticker_data["cash"]) / div_factor
        
        ws.cell(row=row, column=4, value=mc_val).font = input_font
        ws.cell(row=row, column=5, value=ev_val).font = input_font
        ws.cell(row=row, column=6, value=rev_val).font = input_font
        
        growth_cell = ws.cell(row=row, column=7, value=0.10)
        growth_cell.font = input_font
        
        ws.cell(row=row, column=8, value=eb_val).font = input_font
        
        # ゼロ除算保護（IFERROR適用）
        margin_cell = ws.cell(row=row, column=9, value=f"=IFERROR(H{row}/F{row}, 0)")
        margin_cell.font = data_font
        margin_cell.number_format = '0.0%'
        
        ev_rev_cell = ws.cell(row=row, column=10, value=f"=IFERROR(E{row}/F{row}, \"-\")")
        ev_rev_cell.font = data_font
        ev_rev_cell.number_format = '0.00x'
        
        ev_eb_cell = ws.cell(row=row, column=11, value=f"=IFERROR(E{row}/H{row}, \"-\")")
        ev_eb_cell.font = data_font
        ev_eb_cell.number_format = '0.00x'
        
        for col in range(4, 9):
            c = ws.cell(row=row, column=col)
            if col == 7:
                c.number_format = '0.0%'
            else:
                c.number_format = '#,##0.0'
                
        for col_idx in range(1, 12):
            ws.cell(row=row, column=col_idx).border = thin_border
            if "対象会社" in name:
                ws.cell(row=row, column=col_idx).fill = subject_fill
                
    # 統計サマリー
    stats = [
        ("最大値", "MAX"),
        ("75パーセンタイル", "QUARTILE.INC"),
        ("中央値", "MEDIAN"),
        ("25パーセンタイル", "QUARTILE.INC"),
        ("最小値", "MIN")
    ]
    
    stat_start_row = start_row + len(all_data) + 1
    ws.cell(row=stat_start_row - 1, column=1, value="ピア企業統計値").font = Font(name=font_family, size=11, bold=True, color="1B263B")
    
    for idx, (label, func) in enumerate(stats):
        row = stat_start_row + idx
        ws.row_dimensions[row].height = 24
        
        ws.cell(row=row, column=1, value=label).font = bold_data_font
        ws.cell(row=row, column=1).fill = section_fill
        
        peer_range_end = start_row + len(all_data) - 2
        
        for col in range(4, 12):
            col_letter = get_column_letter(col)
            cell = ws.cell(row=row, column=col)
            cell.font = bold_data_font
            cell.border = thin_border
            
            if func == "QUARTILE.INC":
                q_num = 3 if "75" in label else 1
                cell.value = f"=QUARTILE.INC({col_letter}{start_row}:{col_letter}{peer_range_end}, {q_num})"
            else:
                cell.value = f"={func}({col_letter}{start_row}:{col_letter}{peer_range_end})"
                
            if col in [4, 5, 6, 8]:
                cell.number_format = '#,##0.0'
            elif col in [7, 9]:
                cell.number_format = '0.0%'
            elif col in [10, 11]:
                cell.number_format = '0.00x'
                
            if label == "中央値":
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    subject_dir = find_ticker_dir(outdir, ticker_data["ticker"])
    target_path = os.path.join(subject_dir, "analysis")
    os.makedirs(target_path, exist_ok=True)
    out_file = os.path.join(target_path, f"comps_{ticker_data['ticker']}.xlsx")
    wb.save(out_file)
    wb.close()
    logger.info(f"Successfully generated Comps model: {out_file}")

def create_dcf_model(ticker_data, outdir):
    """ディスカウント・キャッシュ・フロー (DCF) シートを構築する"""
    wb = openpyxl.Workbook()
    
    calc_pr = CalcProperties(iterate=True, refMode='A1', iterateCount=100, iterateDelta=0.001)
    wb.properties.calcPr = calc_pr
    
    ws = wb.active
    ws.title = "DCFバリュエーション"
    ws.views.sheetView[0].showGridLines = True
    
    # 共通スタイルインポート
    styles = ExcelStyles()
    
    font_family = styles.font_family
    title_font = styles.title_font
    header_font = styles.header_font
    section_font = styles.section_font
    data_font = styles.data_font
    input_font = styles.input_font
    bold_data_font = styles.bold_data_font
    primary_fill = styles.primary_fill
    section_fill = styles.section_fill
    highlight_fill = styles.highlight_fill
    thin_border_side = styles.thin_border_side
    thin_border = styles.thin_border
    
    is_jpy = ticker_data["currency"] == "JPY"
    unit_str = "十億円" if is_jpy else "百万米ドル"
    div_factor = 1e9 if is_jpy else 1e6
    currency_symbol = "¥" if is_jpy else "$"
    
    # タイトル行
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{ticker_data['name']} ({ticker_data['ticker']}) - ディスカウントキャッシュフロー (DCF) 評価モデル"
    ws["A1"].font = title_font
    ws["A1"].fill = primary_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Inputsシートを作成して前提条件を配置
    ws_inputs = wb.create_sheet(title="前提条件")
    ws_inputs.views.sheetView[0].showGridLines = True

    # タイトル行 (Inputs)
    ws_inputs.merge_cells("A1:C1")
    ws_inputs["A1"] = f"{ticker_data['name']} ({ticker_data['ticker']}) - DCF前提条件"
    ws_inputs["A1"].font = title_font
    ws_inputs["A1"].fill = primary_fill
    ws_inputs["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_inputs.row_dimensions[1].height = 40

    ws_inputs.cell(row=3, column=1, value="項目").font = header_font
    ws_inputs.cell(row=3, column=1).fill = primary_fill
    ws_inputs.cell(row=3, column=2, value="設定値").font = header_font
    ws_inputs.cell(row=3, column=2).fill = primary_fill
    ws_inputs.cell(row=3, column=3, value="情報源 / 備考").font = header_font
    ws_inputs.cell(row=3, column=3).fill = primary_fill
    ws_inputs.row_dimensions[3].height = 25

    rf_rate = 0.010 if is_jpy else 0.040
    tax_rate = 0.306 if is_jpy else 0.210

    inputs_data = [
        ("リスクフリーレート (10年国債利回り)", rf_rate, "0.0%", f"[想定] 10年 {'JGB' if is_jpy else 'US Treasury'} 利回り"),
        ("株式ベータ (対市場ベータ)", 1.20, "0.00", "[想定] 類似企業ベータ"),
        ("株式リスクプレミアム", 0.060, "0.0%", "[想定] 市場リスクプレミアム"),
        ("税引前負債コスト", 0.025 if is_jpy else 0.055, "0.0%", "[想定] 平均負債コスト"),
        ("実効税率", tax_rate, "0.0%", "[想定] 法定実効税率"),
        ("目標負債比率 (負債/総資本)", 0.20, "0.0%", "[想定] 目標資本構成"),
        ("ターミナルEV/EBITDA倍率", 10.0, "0.0x", "[想定] 類似企業マルチプル"),
        ("永久成長率 (ゴードン成長モデル)", 0.005 if is_jpy else 0.020, "0.0%", "[想定] 長期GDP成長率")
    ]

    for idx, (label, val, fmt, src) in enumerate(inputs_data):
        row = 4 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=label).font = data_font
        cell_val = ws_inputs.cell(row=row, column=2, value=val)
        cell_val.font = input_font
        cell_val.number_format = fmt
        cell_val.alignment = Alignment(horizontal="right")
        ws_inputs.cell(row=row, column=3, value=src).font = data_font

        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # Inputsシートの列幅調整
    for col in ws_inputs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_inputs.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # I. 前提条件セクション (WACC & Terminal Multiple)
    ws["A3"] = "I. バリュエーション前提条件"
    ws["A3"].font = section_font
    ws.merge_cells("A3:C3")
    ws["A3"].fill = section_fill
    
    assumptions = [
        ("リスクフリーレート (10年国債利回り)", "=前提条件!B4", "0.0%"),
        ("株式ベータ (対市場ベータ)", "=前提条件!B5", "0.00"),
        ("株式リスクプレミアム", "=前提条件!B6", "0.0%"),
        ("株主資本コスト (CAPM)", "=B4+B5*B6", "0.0%"),
        ("税引前負債コスト", "=前提条件!B7", "0.0%"),
        ("実効税率", "=前提条件!B8", "0.0%"),
        ("税引後負債コスト", "=B8*(1-B9)", "0.0%"),
        ("目標負債比率 (負債/総資本)", "=前提条件!B9", "0.0%"),
        ("目標自己資本比率 (自己資本/総資本)", "=1-B11", "0.0%"),
        ("加重平均資本コスト (WACC)", "=B7*B12+B10*B11", "0.0%"),
        ("ターミナルEV/EBITDA倍率", "=前提条件!B10", "0.0x"),
        ("永久成長率 (ゴードン成長モデル)", "=前提条件!B11", "0.0%")
    ]
    
    for idx, (label, val, fmt) in enumerate(assumptions):
        row = 4 + idx
        ws.row_dimensions[row].height = 20
        ws.cell(row=row, column=1, value=label).font = bold_data_font if ("WACC" in label or "加重平均資本コスト" in label) else data_font
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.font = data_font if str(val).startswith("=") else input_font
        cell_val.number_format = fmt
        cell_val.alignment = Alignment(horizontal="right")
        
        if "WACC" in label or "加重平均資本コスト" in label:
            ws.cell(row=row, column=1).fill = highlight_fill
            cell_val.fill = highlight_fill
            cell_val.font = Font(name=font_family, size=11, bold=True, color="047857")
            
    # II. プロジェクションセクション
    ws["A18"] = f"II. 財務予測 (ベースケース - {unit_str})"
    ws["A18"].font = section_font
    ws.merge_cells("A18:H18")
    ws["A18"].fill = section_fill
    
    proj_headers = ["指標", "実績 (LTM)", "FY1E", "FY2E", "FY3E", "FY4E", "FY5E", "ターミナル"]
    for col_idx, header in enumerate(proj_headers, 1):
        c = ws.cell(row=20, column=col_idx, value=header)
        c.font = header_font
        c.fill = primary_fill
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[20].height = 28
    
    rev_ltm = ticker_data["revenue"] / div_factor if ticker_data["revenue"] else 1000.0
    eb_ltm = ticker_data["ebitda"] / div_factor if ticker_data["ebitda"] else 150.0
    ebit_ltm = ticker_data["ebit"] / div_factor if ticker_data["ebit"] else 100.0
    da_ltm = ticker_data["depreciation"] / div_factor if ticker_data["depreciation"] else 50.0
    capex_ltm = ticker_data["capex"] / div_factor if ticker_data["capex"] else 80.0
    nwc_ltm = ticker_data["nwc_change"] / div_factor if ticker_data["nwc_change"] else 10.0
    
    eb_margin_ltm = eb_ltm / rev_ltm if rev_ltm else 0.15
    da_percent_ltm = da_ltm / rev_ltm if rev_ltm else 0.05
    capex_percent_ltm = capex_ltm / rev_ltm if rev_ltm else 0.08
    nwc_percent_ltm = nwc_ltm / rev_ltm if rev_ltm else 0.01
    
    rows_def = [
        ("売上高", "input", [rev_ltm, "=B21*1.08", "=C21*1.06", "=D21*1.05", "=E21*1.05", "=F21*1.05"], "#,##0.0"),
        ("売上高成長率", "calc", ["", "=(C21-B21)/B21", "=(D21-C21)/C21", "=(E21-D21)/D21", "=(F21-E21)/E21", "=(G21-F21)/F21"], "0.0%"),
        ("EBITDA", "calc", [eb_ltm, f"=C21*{eb_margin_ltm:.4f}", f"=D21*{eb_margin_ltm:.4f}", f"=E21*{eb_margin_ltm:.4f}", f"=F21*{eb_margin_ltm:.4f}", f"=G21*{eb_margin_ltm:.4f}"], "#,##0.0"),
        ("EBITDAマージン", "calc", ["=B23/B21", "=C23/C21", "=D23/D21", "=E23/E21", "=F23/F21", "=G23/G21"], "0.0%"),
        ("減価償却費", "calc", [da_ltm, f"=C21*{da_percent_ltm:.4f}", f"=D21*{da_percent_ltm:.4f}", f"=E21*{da_percent_ltm:.4f}", f"=F21*{da_percent_ltm:.4f}", f"=G21*{da_percent_ltm:.4f}"], "#,##0.0"),
        ("営業利益 (EBIT)", "calc", ["=B23-B25", "=C23-C25", "=D23-D25", "=E23-E25", "=F23-F25", "=G23-G25"], "#,##0.0"),
        ("営業利益に対する税金", "calc", ["", "=C26*$B$9", "=D26*$B$9", "=E26*$B$9", "=F26*$B$9", "=G26*$B$9"], "#,##0.0"),
        ("税引後営業利益 (NOPAT)", "calc", ["", "=C26-C27", "=D26-D27", "=E26-E27", "=F26-F27", "=G26-G27"], "#,##0.0"),
        ("加算：減価償却費", "calc", ["", "=C25", "=D25", "=E25", "=F25", "=G25"], "#,##0.0"),
        ("減算：設備投資額 (CapEx)", "calc", [capex_ltm, f"=C21*{capex_percent_ltm:.4f}", f"=D21*{capex_percent_ltm:.4f}", f"=E21*{capex_percent_ltm:.4f}", f"=F21*{capex_percent_ltm:.4f}", f"=G21*{capex_percent_ltm:.4f}"], "#,##0.0"),
        ("減算：運転資本増減額", "calc", [nwc_ltm, f"=C21*{nwc_percent_ltm:.4f}", f"=D21*{nwc_percent_ltm:.4f}", f"=E21*{nwc_percent_ltm:.4f}", f"=F21*{nwc_percent_ltm:.4f}", f"=G21*{nwc_percent_ltm:.4f}"], "#,##0.0"),
        ("企業フリーキャッシュフロー (FCFF)", "calc", ["", "=C28+C29-C30-C31", "=D28+D29-D30-D31", "=E28+E29-E30-E31", "=F28+F29-F30-F31", "=G28+G29-G30-G31"], "#,##0.0"),
        ("割引期間", "calc", ["", 0.5, 1.5, 2.5, 3.5, 4.5], "0.0"),
        ("割引因子", "calc", ["", "=1/((1+$B$13)^C33)", "=1/((1+$B$13)^D33)", "=1/((1+$B$13)^E33)", "=1/((1+$B$13)^F33)", "=1/((1+$B$13)^G33)"], "0.0000"),
        ("FCFFの現在価値", "calc", ["", "=C32*C34", "=D32*D34", "=E32*E34", "=F32*F34", "=G32*G34"], "#,##0.0")
    ]
    
    for r_idx, (label, r_type, vals, fmt) in enumerate(rows_def):
        row = 21 + r_idx
        ws.row_dimensions[row].height = 22
        ws.cell(row=row, column=1, value=label).font = bold_data_font if ("FCFF" in label or "現在価値" in label or "フリーキャッシュフロー" in label) else data_font
        
        if "FCFF" in label or "現在価値" in label or "フリーキャッシュフロー" in label:
            for col_idx in range(1, 9):
                ws.cell(row=row, column=col_idx).border = Border(top=thin_border_side, bottom=thin_border_side)
                ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FAF5FF", end_color="FAF5FF", fill_type="solid")
                
        for c_idx, val in enumerate(vals):
            col = 2 + c_idx
            cell = ws.cell(row=row, column=col)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            
            if str(val).startswith("="):
                cell.value = val
                cell.font = data_font
            else:
                cell.value = val
                cell.font = input_font if (r_type == "input" and col == 2) else data_font
                
    ws.cell(row=23, column=8, value="=G23*(1+$B$15)").font = data_font
    ws.cell(row=23, column=8).number_format = "#,##0.0"
    
    # III. 企業価値ブリッジ
    ws["A38"] = "III. 企業価値評価ブリッジおよび結論"
    ws["A38"].font = section_font
    ws.merge_cells("A38:C38")
    ws["A38"].fill = section_fill
    
    debt_val = ticker_data["total_debt"] / div_factor if ticker_data["total_debt"] else 100.0
    cash_val = ticker_data["cash"] / div_factor if ticker_data["cash"] else 50.0
    shares_outstanding_m = ticker_data["shares_outstanding"] / 1e6 if ticker_data["shares_outstanding"] else 100.0
    current_price_raw = ticker_data["current_price"] if ticker_data["current_price"] else 100.0
    
    bridge = [
        ("予測期間FCFFの現在価値累計", "=SUM(C35:G35)", "#,##0.0"),
        ("ターミナルバリュー (マルチプル法)", "=H23*$B$14", "#,##0.0"),
        ("ターミナルバリューの現在価値", "=B40*G34", "#,##0.0"),
        ("事業価値 (EV)", "=B39+B41", "#,##0.0"),
        ("控除：有利子負債合計", debt_val, "#,##0.0"),
        ("加算：現金及び現金同等物", cash_val, "#,##0.0"),
        ("株式価値", "=B42-B43+B44", "#,##0.0"),
        ("発行済株式数 (百万株)", shares_outstanding_m, "#,##0.0"),
        ("理論株価", "=(B45/B46)*100" if is_jpy else "=B45/B46", f"{currency_symbol}#,##0.00"),
        ("現在株価", current_price_raw, f"{currency_symbol}#,##0.00"),
        ("理論プレミアム / (ディスカウント)", "=(B47-B48)/B48", "0.0%")
    ]
    
    for idx, (label, val, fmt) in enumerate(bridge):
        row = 39 + idx
        ws.row_dimensions[row].height = 20
        ws.cell(row=row, column=1, value=label).font = bold_data_font if ("株価" in label or "EV" in label or "事業価値" in label or "株式価値" in label) else data_font
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.font = data_font if str(val).startswith("=") else input_font
        cell_val.number_format = fmt
        cell_val.alignment = Alignment(horizontal="right")
        
        if "株価" in label:
            ws.cell(row=row, column=1).fill = highlight_fill
            cell_val.fill = highlight_fill
            cell_val.font = Font(name=font_family, size=11, bold=True, color="047857")
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
        
    subject_dir = find_ticker_dir(outdir, ticker_data["ticker"])
    target_path = os.path.join(subject_dir, "analysis")
    os.makedirs(target_path, exist_ok=True)
    out_file = os.path.join(target_path, f"dcf_{ticker_data['ticker']}.xlsx")
    wb.save(out_file)
    wb.close()
    logger.info(f"Successfully generated DCF model: {out_file}")

def main():
    args = parse_args()
    ticker_str = normalize_ticker(args.ticker)
    
    logger.info(f"Running generic financial modeling for {ticker_str}...")
    ticker_dir = find_ticker_dir(args.outdir, ticker_str)
    
    if not os.path.exists(ticker_dir):
        logger.error(f"Directory {ticker_dir} does not exist. Please run fetch_yfinance.py first.")
        sys.exit(1)
        
    try:
        ticker_data = get_latest_financial_data(ticker_dir, ticker_str)
    except FileNotFoundError as e:
        logger.error(str(e))
        sys.exit(1)
        
    peers = []
    if args.peers:
        peers = [normalize_ticker(p) for p in args.peers.split(",") if p.strip()]
        
    create_dcf_model(ticker_data, args.outdir)
    
    if peers:
        create_comps_model(ticker_data, peers, args.outdir)
    else:
        logger.info("Skipped Comps model generation (no peers specified).")

if __name__ == "__main__":
    main()
