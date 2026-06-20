import os
import sys
import glob
import json
import logging
from openpyxl.styles import Font, PatternFill, Border, Side

# pandas のインポートチェック
try:
    import pandas as pd
except ImportError:
    print("Error: pandas is required. Please install it in the virtual environment.", file=sys.stderr)
    sys.exit(1)

def setup_logging(name=None):
    """標準的なロギングを設定する"""
    logger = logging.getLogger(name or "financial_services")
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        handler = logging.StreamHandler(sys.stdout)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
    return logger

logger = setup_logging()

def normalize_ticker(ticker_str):
    """日本株の4桁の数字ティッカーに自動で '.T' を付与して正規化する"""
    if not ticker_str:
        return ""
    ticker_str = ticker_str.strip()
    if len(ticker_str) == 4 and ticker_str[0].isdigit() and ticker_str.isalnum():
        return f"{ticker_str}.T"
    return ticker_str

def find_ticker_dir(base_dir, ticker_str):
    """base_dir配下から、ticker_strで始まるティッカーフォルダを動的に探索する"""
    normalized = normalize_ticker(ticker_str)
    pattern = os.path.join(base_dir, f"{normalized}*")
    matches = glob.glob(pattern)
    dirs = [m for m in matches if os.path.isdir(m)]
    # 部分一致で意図しないフォルダが選ばれないよう、ティッカーコード直後に '_' またはマッチするものを厳格にソート
    # (例: 285A.T_Kioxia_Holdings が優先されるようにする)
    dirs.sort(key=len, reverse=True)
    if dirs:
        return dirs[0]
    return os.path.join(base_dir, normalized)

def get_latest_financial_data(ticker_dir, ticker_str):
    """CSVおよびJSONから実績データを読み込み、辞書として返す"""
    normalized = normalize_ticker(ticker_str)
    
    # 必須ファイルの存在確認。存在しない場合は例外を発生させる
    inc_path = os.path.join(ticker_dir, "market_data", "annual_income_stmt.csv")
    bs_path = os.path.join(ticker_dir, "market_data", "annual_balance_sheet.csv")
    summary_path = os.path.join(ticker_dir, "market_data", "summary.json")
    
    if not os.path.exists(summary_path) or not os.path.exists(inc_path) or not os.path.exists(bs_path):
        raise FileNotFoundError(f"Required financial data is missing in {ticker_dir}. Please run fetch_yfinance.py first.")

    data = {
        "ticker": normalized,
        "name": normalized,
        "currency": "JPY" if normalized.endswith(".T") else "USD",
        "current_price": 0.0,
        "shares_outstanding": 0.0,
        "market_cap": 0.0,
        "revenue": 0.0,
        "ebitda": 0.0,
        "ebit": 0.0,
        "depreciation": 0.0,
        "tax_provision": 0.0,
        "capex": 0.0,
        "nwc_change": 0.0,
        "total_debt": 0.0,
        "cash": 0.0
    }
    
    # 1. summary.json の読み込み
    with open(summary_path, "r", encoding="utf-8") as f:
        summary = json.load(f)
        data["name"] = summary.get("long_name") or summary.get("ticker")
        data["currency"] = summary.get("currency") or data["currency"]
        data["current_price"] = summary.get("current_price") or 0.0
        data["shares_outstanding"] = summary.get("shares_outstanding") or 0.0
        data["market_cap"] = summary.get("market_cap") or 0.0
        data["ebitda"] = summary.get("ebitda") or 0.0
        
        # 日本株の100倍バグ補正
        # 株価・時価総額が異常値（100倍）で取得されている場合に補正
        if normalized.endswith(".T") and data["current_price"] > 50000:
            # 補正前ログ
            logger.info(f"Detected potential 100x pricing bug for {normalized}. Raw price: {data['current_price']}. Applying correction.")
            data["current_price"] /= 100
            if data["market_cap"] > 1e13: # 10兆円以上になっている場合
                data["market_cap"] /= 100
                
    # ヘルパー関数: 安全に値を取得する
    def get_val(df, index_names, col):
        for name in index_names:
            if name in df.index:
                val = df.loc[name, col]
                if isinstance(val, pd.Series):
                    val = val.iloc[0]
                if pd.notna(val):
                    return float(val)
        return 0.0

    # 有効なデータが入っている最初の列を探すヘルパー
    def get_latest_actual_col(df, index_names):
        for name in index_names:
            if name in df.index:
                for col in df.columns:
                    val = df.loc[name, col]
                    if isinstance(val, pd.Series):
                        val = val.iloc[0]
                    if pd.notna(val) and val != 0:
                        return col
        return df.columns[0]

    # 2. annual_income_stmt.csv
    df_inc = pd.read_csv(inc_path, index_col=0)
    latest_col_inc = get_latest_actual_col(df_inc, ["Total Revenue", "Operating Revenue", "Revenue"])
    
    data["revenue"] = get_val(df_inc, ["Total Revenue", "Operating Revenue", "Revenue"], latest_col_inc)
    ebitda_val = get_val(df_inc, ["EBITDA", "Normalized EBITDA"], latest_col_inc)
    ebit_val = get_val(df_inc, ["EBIT", "Operating Income", "Total Operating Income As Reported"], latest_col_inc)
    dep_val = get_val(df_inc, ["Reconciled Depreciation", "Depreciation And Amortization", "Depreciation"], latest_col_inc)
    
    data["ebit"] = ebit_val
    data["depreciation"] = dep_val
    data["ebitda"] = ebitda_val if ebitda_val != 0 else (ebit_val + dep_val)
    data["tax_provision"] = get_val(df_inc, ["Tax Provision", "Income Tax Expense"], latest_col_inc)
    
    # 3. annual_balance_sheet.csv
    df_bs = pd.read_csv(bs_path, index_col=0)
    latest_col_bs = get_latest_actual_col(df_bs, ["Total Assets"])
    data["total_debt"] = get_val(df_bs, ["Total Debt", "Long Term Debt", "Current Debt"], latest_col_bs)
    data["cash"] = get_val(df_bs, ["Cash Cash Equivalents And Short Term Investments", "Cash And Cash Equivalents", "Cash"], latest_col_bs)
    
    # 4. annual_cashflow.csv
    cf_path = os.path.join(ticker_dir, "market_data", "annual_cashflow.csv")
    if os.path.exists(cf_path):
        df_cf = pd.read_csv(cf_path, index_col=0)
        latest_col_cf = get_latest_actual_col(df_cf, ["Operating Cash Flow", "Net Income From Continuing Operations"])
        data["capex"] = abs(get_val(df_cf, ["Capital Expenditure", "Purchase Of PPE", "Net PPE Purchase And Sale"], latest_col_cf))
        data["nwc_change"] = get_val(df_cf, ["Change In Working Capital", "Changes In Cash"], latest_col_cf)
        
    return data

class ExcelStyles:
    def __init__(self, font_family="Outfit"):
        self.font_family = font_family
        self.title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
        self.header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        self.section_font = Font(name=font_family, size=12, bold=True, color="1B263B")
        self.data_font = Font(name=font_family, size=11, color="000000")
        self.input_font = Font(name=font_family, size=11, color="003366")
        self.bold_data_font = Font(name=font_family, size=11, bold=True, color="000000")
        
        # 配色
        self.primary_fill = PatternFill(start_color="1B263B", end_color="1B263B", fill_type="solid")
        self.section_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        self.subject_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # 罫線
        self.thin_border_side = Side(style='thin', color='CBD5E1')
        self.thin_border = Border(left=self.thin_border_side, right=self.thin_border_side, top=self.thin_border_side, bottom=self.thin_border_side)
