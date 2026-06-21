import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import argparse
from utils import find_ticker_dir, get_latest_financial_data, ExcelStyles, setup_logging

logger = setup_logging("generate_lbo")

def build_lbo_model(ticker_data, ticker_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LBOバリュエーション"
    ws.views.sheetView[0].showGridLines = True
    
    # 共通スタイルインポート
    styles = ExcelStyles()
    
    font_family = styles.font_family
    title_font = styles.title_font
    header_font = styles.header_font
    section_font = styles.section_font
    data_font = styles.data_font
    input_font = styles.input_font
    bold_data_font = styles.bold_data_font
    primary_fill = styles.primary_fill
    section_fill = styles.section_fill
    highlight_fill = styles.highlight_fill
    thin_border_side = styles.thin_border_side
    thin_border = styles.thin_border
    
    is_jpy = ticker_data["currency"] == "JPY"
    unit_str = "十億円" if is_jpy else "百万ドル"
    div_factor = 1e9 if is_jpy else 1e6
    
    # 実績値
    mc_val = ticker_data["market_cap"] / div_factor if ticker_data["market_cap"] else 593.0
    debt_val = ticker_data["total_debt"] / div_factor if ticker_data["total_debt"] else 439.5
    cash_val = ticker_data["cash"] / div_factor if ticker_data["cash"] else 167.9
    eb_val = ticker_data["ebitda"] / div_factor if ticker_data["ebitda"] else 768.3
    da_val = ticker_data["depreciation"] / div_factor if ticker_data["depreciation"] else eb_val * 0.4
    
    # タイトル
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{ticker_data['name']} ({ticker_data['ticker']}) - レバレッジド・バイアウト（LBO）バリュエーションモデル"
    ws["A1"].font = title_font
    ws["A1"].fill = primary_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 前提条件シートを作成して配置
    ws_inputs = wb.create_sheet(title="前提条件")
    ws_inputs.views.sheetView[0].showGridLines = True

    # タイトル行 (前提条件)
    ws_inputs.merge_cells("A1:C1")
    ws_inputs["A1"] = f"{ticker_data['name']} ({ticker_data['ticker']}) - LBOバリュエーション前提条件"
    ws_inputs["A1"].font = title_font
    ws_inputs["A1"].fill = primary_fill
    ws_inputs["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_inputs.row_dimensions[1].height = 40

    ws_inputs.cell(row=3, column=1, value="項目").font = header_font
    ws_inputs.cell(row=3, column=1).fill = primary_fill
    ws_inputs.cell(row=3, column=2, value="値").font = header_font
    ws_inputs.cell(row=3, column=2).fill = primary_fill
    ws_inputs.cell(row=3, column=3, value="出典 / メモ").font = header_font
    ws_inputs.cell(row=3, column=3).fill = primary_fill
    ws_inputs.row_dimensions[3].height = 25

    inputs_data = [
        ("買収プレミアム", 0.30, "0.0%", "[前提条件] 買収プレミアム"),
        ("取引費用・諸経費", 15.0, "#,##0.0", "[前提条件] 取引費用"),
        ("LBO負債調達比率 (TEV比)", 0.60, "0.0%", "[前提条件] TEVに対する負債割合"),
        ("シニアローン LTM EBITDA倍率", 4.0, "0.0x", "[前提条件] シニアローンレバレッジ倍率"),
        ("シニアローン金利", 0.06, "0.0%", "[前提条件] シニアローン金利"),
        ("メザニン金利", 0.10, "0.0%", "[前提条件] メザニン金利"),
        ("実効税率", 0.306, "0.0%", "[前提条件] 実効税率"),
    ]

    for idx, (label, val, fmt, src) in enumerate(inputs_data):
        row = 4 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=label).font = data_font
        cell_val = ws_inputs.cell(row=row, column=2, value=val)
        cell_val.font = input_font
        cell_val.number_format = fmt
        cell_val.alignment = Alignment(horizontal="right")
        ws_inputs.cell(row=row, column=3, value=src).font = data_font
        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # 設備投資額 (CapEx) 予測
    ws_inputs.cell(row=12, column=1, value="設備投資額 (CapEx) 予測").font = bold_data_font
    ws_inputs.cell(row=13, column=1, value="年度").font = header_font
    ws_inputs.cell(row=13, column=1).fill = primary_fill
    ws_inputs.cell(row=13, column=2, value="設備投資額").font = header_font
    ws_inputs.cell(row=13, column=2).fill = primary_fill
    ws_inputs.cell(row=13, column=3, value="出典 / メモ").font = header_font
    ws_inputs.cell(row=13, column=3).fill = primary_fill
    ws_inputs.row_dimensions[13].height = 20

    capex_inputs = [
        ("FY25E", -150.0, "[前提条件]"),
        ("FY26E", -160.0, "[前提条件]"),
        ("FY27E", -170.0, "[前提条件]"),
        ("FY28E", -170.0, "[前提条件]"),
        ("FY29E", -180.0, "[前提条件]"),
    ]
    for idx, (yr, val, src) in enumerate(capex_inputs):
        row = 14 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=yr).font = data_font
        cell_val = ws_inputs.cell(row=row, column=2, value=val)
        cell_val.font = input_font
        cell_val.number_format = "#,##0.0"
        cell_val.alignment = Alignment(horizontal="right")
        ws_inputs.cell(row=row, column=3, value=src).font = data_font
        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # 運転資本増減額予測
    ws_inputs.cell(row=20, column=1, value="運転資本増減額予測").font = bold_data_font
    ws_inputs.cell(row=21, column=1, value="年度").font = header_font
    ws_inputs.cell(row=21, column=1).fill = primary_fill
    ws_inputs.cell(row=21, column=2, value="運転資本増減額").font = header_font
    ws_inputs.cell(row=21, column=2).fill = primary_fill
    ws_inputs.cell(row=21, column=3, value="出典 / メモ").font = header_font
    ws_inputs.cell(row=21, column=3).fill = primary_fill
    ws_inputs.row_dimensions[21].height = 20

    nwc_inputs = [
        ("FY25E", -10.0, "[前提条件]"),
        ("FY26E", -12.0, "[前提条件]"),
        ("FY27E", -15.0, "[前提条件]"),
        ("FY28E", -15.0, "[前提条件]"),
        ("FY29E", -15.0, "[前提条件]"),
    ]
    for idx, (yr, val, src) in enumerate(nwc_inputs):
        row = 22 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=yr).font = data_font
        cell_val = ws_inputs.cell(row=row, column=2, value=val)
        cell_val.font = input_font
        cell_val.number_format = "#,##0.0"
        cell_val.alignment = Alignment(horizontal="right")
        ws_inputs.cell(row=row, column=3, value=src).font = data_font
        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # エグジットマルチプル
    ws_inputs.cell(row=28, column=1, value="エグジットマルチプル").font = bold_data_font
    ws_inputs.cell(row=29, column=1, value="ケース").font = header_font
    ws_inputs.cell(row=29, column=1).fill = primary_fill
    ws_inputs.cell(row=29, column=2, value="エグジットマルチプル").font = header_font
    ws_inputs.cell(row=29, column=2).fill = primary_fill
    ws_inputs.cell(row=29, column=3, value="出典 / メモ").font = header_font
    ws_inputs.cell(row=29, column=3).fill = primary_fill
    ws_inputs.row_dimensions[29].height = 20

    exit_inputs = [
        ("ダウンサイド", 8.0, "[前提条件]"),
        ("ベース", 10.0, "[前提条件]"),
        ("アップサイド", 12.0, "[前提条件]"),
    ]
    for idx, (case, val, src) in enumerate(exit_inputs):
        row = 30 + idx
        ws_inputs.row_dimensions[row].height = 20
        ws_inputs.cell(row=row, column=1, value=case).font = data_font
        cell_val = ws_inputs.cell(row=row, column=2, value=val)
        cell_val.font = input_font
        cell_val.number_format = "0.0x"
        cell_val.alignment = Alignment(horizontal="right")
        ws_inputs.cell(row=row, column=3, value=src).font = data_font
        for col in range(1, 4):
            ws_inputs.cell(row=row, column=col).border = thin_border

    # Inputsシートの列幅調整
    for col in ws_inputs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_inputs.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # I. 取引前提条件
    ws["A3"] = "I. 取引前提条件"
    ws["A3"].font = section_font
    ws.merge_cells("A3:C3")
    ws["A3"].fill = section_fill
    
    tx_assumptions = [
        ("対象企業株式価値 (時価総額)", mc_val, "#,##0.0"),
        ("買収プレミアム", "=前提条件!B4", "0.0%"),
        ("買収提案株式価値", "=B4*(1+B5)", "#,##0.0"),
        ("既存純有利子負債 (借換対象)", debt_val - cash_val, "#,##0.0"),
        ("取引費用・諸経費", "=前提条件!B5", "#,##0.0"),
        ("企業価値合計 (TEV)", "=B6+B7+B8", "#,##0.0"),
        ("LBO負債調達比率 (TEV比)", "=前提条件!B6", "0.0%"),
        ("シニアローン LTM EBITDA倍率", "=前提条件!B7", "0.0x"),
        ("必要スポンサー出資額", "=B9*(1-B10)", "#,##0.0")
    ]
    
    for idx, (label, val, fmt) in enumerate(tx_assumptions):
        row = 4 + idx
        ws.row_dimensions[row].height = 20
        ws.cell(row=row, column=1, value=label).font = bold_data_font if ("株式価値" in label or "TEV" in label or "出資額" in label) else data_font
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.font = data_font if str(val).startswith("=") else input_font
        cell_val.number_format = fmt
        cell_val.alignment = Alignment(horizontal="right")
        
    # II. 資金の調達と使途 (Sources & Uses)
    ws["A15"] = "II. 資金の調達と使途 (Sources & Uses)"
    ws["A15"].font = section_font
    ws.merge_cells("A15:G15")
    ws["A15"].fill = section_fill
    
    ws["A17"] = "資金の調達 (Sources)"
    ws["A17"].font = bold_data_font
    ws["B17"] = "金額"
    ws["B17"].font = bold_data_font
    ws["C17"] = "構成比"
    ws["C17"].font = bold_data_font
    
    ws["A18"] = "シニアローン (4.0x EBITDA)"
    ws["B18"] = f"=B11*{eb_val}"
    ws["C18"] = "=B18/$B$21"
    
    ws["A19"] = "メザニン資金調達"
    ws["B19"] = "=B9*B10-B18"
    ws["C19"] = "=B19/$B$21"
    
    ws["A20"] = "スポンサー出資金"
    ws["B20"] = "=B12"
    ws["C20"] = "=B20/$B$21"
    
    ws["A21"] = "調達合計"
    ws["B21"] = "=SUM(B18:B20)"
    ws["C21"] = "=SUM(C18:C20)"
    
    ws["E17"] = "資金の使途 (Uses)"
    ws["E17"].font = bold_data_font
    ws["F17"] = "金額"
    ws["F17"].font = bold_data_font
    ws["G17"] = "構成比"
    ws["G17"].font = bold_data_font
    
    ws["E18"] = "対象企業株式の買収"
    ws["F18"] = "=B6"
    ws["G18"] = "=F18/$F$21"
    
    ws["E19"] = "既存負債の借換"
    ws["F19"] = "=B7"
    ws["G19"] = "=F19/$F$21"
    
    ws["E20"] = "取引費用・諸経費"
    ws["F20"] = "=B8"
    ws["G20"] = "=F20/$F$21"
    
    ws["E21"] = "使途合計"
    ws["F21"] = "=SUM(F18:F20)"
    ws["G21"] = "=SUM(G18:G20)"
    
    # 書式設定
    for r in range(18, 22):
        for c in [2, 3, 6, 7]:
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="right")
            if c in [3, 7]:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0.0'
            if r == 21:
                cell.font = bold_data_font
                cell.border = Border(top=thin_border_side, bottom=Side(style='double', color='1B263B'))
                
    # III. キャッシュ・フローおよび債務返済シミュレーション ({unit_str})
    ws["A24"] = f"III. キャッシュ・フローおよび債務返済シミュレーション ({unit_str})"
    ws["A24"].font = section_font
    ws.merge_cells("A24:G24")
    ws["A24"].fill = section_fill
    
    proj_headers = ["項目", "FY25E", "FY26E", "FY27E", "FY28E", "FY29E"]
    for col_idx, header in enumerate(proj_headers, 1):
        c = ws.cell(row=26, column=col_idx, value=header)
        c.font = header_font
        c.fill = primary_fill
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[26].height = 24
    
    cf_data = [
        ("EBITDA", [eb_val, "=B27*1.08", "=C27*1.06", "=D27*1.05", "=E27*1.05"]),
        ("減算：減価償却費", [da_val, "=B28*1.05", "=C28*1.05", "=D28*1.04", "=E28*1.04"]),
        ("営業利益 (EBIT)", ["=B27-B28", "=C27-C28", "=D27-D28", "=E27-E28", "=F27-F28"]),
        ("減算：支払利息 (シニア + メザニン)", ["=B38*前提条件!$B$8+$B$19*前提条件!$B$9", "=C38*前提条件!$B$8+$B$19*前提条件!$B$9", "=D38*前提条件!$B$8+$B$19*前提条件!$B$9", "=E38*前提条件!$B$8+$B$19*前提条件!$B$9", "=F38*前提条件!$B$8+$B$19*前提条件!$B$9"]),
        ("税引前当期純利益", ["=B29-B30", "=C29-C30", "=D29-D30", "=E29-E30", "=F29-F30"]),
        ("減算：法人税等 (30.6%)", ["=B31*前提条件!$B$10", "=C31*前提条件!$B$10", "=D31*前提条件!$B$10", "=E31*前提条件!$B$10", "=F31*前提条件!$B$10"]),
        ("当期純利益", ["=B31-B32", "=C31-C32", "=D31-D32", "=E31-E32", "=F31-F32"]),
        ("加算：減価償却費", ["=B28", "=C28", "=D28", "=E28", "=F28"]),
        ("減算：設備投資額", ["=前提条件!B14", "=前提条件!B15", "=前提条件!B16", "=前提条件!B17", "=前提条件!B18"]),
        ("減算：運転資本増減額", ["=前提条件!B22", "=前提条件!B23", "=前提条件!B24", "=前提条件!B25", "=前提条件!B26"]),
        ("フリーキャッシュフロー (債務返済原資)", ["=B33+B34+B35+B36", "=C33+C34+C35+C36", "=D33+D34+D35+D36", "=E33+E34+E35+E36", "=F33+F34+F35+F36"]),
        
        ("シニアローン期首残高", ["=B18", "=B40", "=C40", "=D40", "=E40"]),
        ("減算：債務返済 (FCF全額回収)", ["=MIN(B38, B37)", "=MIN(C38, C37)", "=MIN(D38, D37)", "=MIN(E38, E37)", "=MIN(F38, F37)"]),
        ("シニアローン期末残高", ["=B38-B39", "=C38-C39", "=D38-D39", "=E38-E39", "=F38-F39"])
    ]
    
    for idx, (label, vals) in enumerate(cf_data):
        row = 27 + idx
        ws.row_dimensions[row].height = 20
        ws.cell(row=row, column=1, value=label).font = bold_data_font if ("残高" in label or "フリーキャッシュフロー" in label) else data_font
        for col_idx, val in enumerate(vals, 2):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(horizontal="right")
            cell.value = val
            cell.number_format = '#,##0.0'
            if ("残高" in label or "フリーキャッシュフロー" in label):
                cell.fill = section_fill
                
    # IV. スポンサー投資リターン分析 (5年後エグジット)
    ws["A43"] = "IV. スポンサー投資リターン分析 (5年後エグジット)"
    ws["A43"].font = section_font
    ws.merge_cells("A43:G43")
    ws["A43"].fill = section_fill
    
    ws["A45"] = "エグジットEV/EBITDA倍率"
    ws["B45"] = "エグジットEV"
    ws["C45"] = "控除：純有利子負債"
    ws["D45"] = "エグジット時株式価値"
    ws["E45"] = "スポンサーリターン倍率 (MoIC)"
    ws["F45"] = "スポンサーIRR"
    
    for col in range(1, 7):
        ws.cell(row=45, column=col).font = bold_data_font
        
    exit_multiples_refs = ["=前提条件!B30", "=前提条件!B31", "=前提条件!B32"]
    for idx, ref in enumerate(exit_multiples_refs):
        row = 46 + idx
        ws.row_dimensions[row].height = 22
        
        cell_mult = ws.cell(row=row, column=1, value=ref)
        cell_mult.number_format = '0.0x'
        cell_mult.font = data_font
        
        # Exit EV = Exit Multiple * FY29 EBITDA (F27)
        ws.cell(row=row, column=2, value=f"=A{row}*F27").number_format = '#,##0.0'
        # Less: Net Debt = Ending Senior Debt (F41) + Mezzanine (B19)
        ws.cell(row=row, column=3, value=f"=F41+$B$19").number_format = '#,##0.0'
        # Exit Equity Value = EV - Net Debt
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = '#,##0.0'
        # MoIC = Exit Equity Value / Initial Sponsor Equity (B20)
        ws.cell(row=row, column=5, value=f"=D{row}/$B$20").number_format = '0.00x'
        # IRR = (MoIC)^(1/5) - 1
        ws.cell(row=row, column=6, value=f"=(E{row})^(1/5)-1").number_format = '0.0%'
        
        # ベースケース（インデックス1）をハイライト
        if idx == 1:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = highlight_fill
                ws.cell(row=row, column=col).font = bold_data_font
                
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
            
    # 列幅調整
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
        
    target_path = os.path.join(ticker_dir, "analysis")
    os.makedirs(target_path, exist_ok=True)
    out_file = os.path.join(target_path, f"lbo_{ticker_data['ticker']}.xlsx")
    
    # 循環参照解決のための反復計算設定を有効化
    from openpyxl.workbook.properties import CalcProperties
    calc_pr = CalcProperties(iterate=True, refMode='A1', iterateCount=100, iterateDelta=0.001)
    wb.properties.calcPr = calc_pr
    
    wb.save(out_file)
    wb.close()
    logger.info(f"Successfully generated LBO model: {out_file}")

def main():
    parser = argparse.ArgumentParser(description="Generate LBO valuation model")
    parser.add_argument("ticker", type=str, help="Stock ticker")
    parser.add_argument("--outdir", type=str, default="./out", help="Base output directory")
    args = parser.parse_args()
    
    ticker_str = args.ticker.strip()
    ticker_dir = find_ticker_dir(args.outdir, ticker_str)
    
    try:
        ticker_data = get_latest_financial_data(ticker_dir, ticker_str)
    except FileNotFoundError as e:
        logger.error(str(e))
        sys.exit(1)
        
    build_lbo_model(ticker_data, ticker_dir)

if __name__ == "__main__":
    main()
